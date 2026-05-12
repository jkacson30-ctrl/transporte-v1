# =============================================================================
# MINEBALANCE — Streamlit Web App
# Net Smelter Return · Monte Carlo · Factor de Corrección Adaptativo
# Universidad Nacional del Altiplano — Ingeniería de Minas — 9no Semestre 2026
# Autor: Machaca Espinoza, Jkacson Ruso
# =============================================================================

import streamlit as st
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import warnings
import io
import datetime

warnings.filterwarnings("ignore")

# ── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="MineBalance",
    page_icon="⛏️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Global colors ──────────────────────────────────────────────────────────
COLOR_CYAN   = "#00C6FF"
COLOR_HEAP   = "#FFB347"
COLOR_DUMP   = "#FF6B6B"
COLOR_GREEN  = "#00E676"
COLOR_YELLOW = "#FFD600"
COLOR_RED    = "#D50000"
COLOR_ACCENT = "#7C4DFF"

STYLE = {
    "figure.facecolor": "#0E1117",
    "axes.facecolor":   "#1E2130",
    "axes.edgecolor":   "#3A3F58",
    "axes.labelcolor":  "#C9D1E0",
    "axes.titlecolor":  "#FFFFFF",
    "xtick.color":      "#C9D1E0",
    "ytick.color":      "#C9D1E0",
    "text.color":       "#C9D1E0",
    "grid.color":       "#2E3450",
    "grid.linestyle":   "--",
    "grid.alpha":       0.5,
    "legend.facecolor": "#1E2130",
    "legend.edgecolor": "#3A3F58",
}
plt.rcParams.update(STYLE)

EXCEL_HEADER_FILL = "1F4E79"
EXCEL_GREEN_FILL  = "C6EFCE"
EXCEL_YELLOW_FILL = "FFEB9C"
EXCEL_RED_FILL    = "FFC7CE"

# =============================================================================
# SECTION 1 — CALCULATION FUNCTIONS
# =============================================================================

def calc_corrected_grades(grade_au, grade_ag, fc_lab):
    return grade_au * fc_lab, grade_ag * fc_lab


def calc_nsr(tonnage, grade_au, grade_ag, rec_au, rec_ag,
             price_au, price_ag, deduction_frac, cost_process, cost_transport):
    gross_au   = tonnage * grade_au * rec_au * price_au
    gross_ag   = tonnage * grade_ag * rec_ag * price_ag
    net_income = (gross_au + gross_ag) * (1 - deduction_frac)
    total_cost = tonnage * (cost_process + cost_transport)
    nsr        = net_income - total_cost
    return nsr, gross_au, gross_ag, net_income, total_cost


def calc_decision(nsr_cian, nsr_heap, grade_au_corr, cutoff_cian, cutoff_heap):
    if nsr_cian > nsr_heap and grade_au_corr >= cutoff_cian:
        return "CIANURACIÓN", nsr_cian
    elif nsr_heap > 0 and grade_au_corr >= cutoff_heap:
        return "HEAP LEACH", nsr_heap
    else:
        return "BOTADERO", 0.0


def monte_carlo(p):
    n   = p["n_simulations"]
    df  = p["deduction"] / 100.0
    T   = p["tonnage"]
    gau, gag = calc_corrected_grades(p["grade_au"], p["grade_ag"], p["fc_lab"])
    sig_ag   = p["grade_ag"] * 0.15

    rng = np.random.default_rng(2026)
    l_au = np.maximum(0, rng.normal(gau, p["sigma_m"], n))
    l_ag = np.maximum(0, rng.normal(gag, sig_ag, n))

    rec_au_c = np.maximum(0.50, rng.normal(p["rec_au_cian"]/100, 0.03, n))
    rec_ag_c = np.maximum(0.40, rng.normal(p["rec_ag_cian"]/100, 0.04, n))
    rec_au_h = np.maximum(0.30, rng.normal(p["rec_au_heap"]/100, 0.05, n))
    rec_ag_h = np.maximum(0.10, rng.normal(p["rec_ag_heap"]/100, 0.06, n))

    nsr_c = ((T*l_au*rec_au_c*p["price_au"] + T*l_ag*rec_ag_c*p["price_ag"])*(1-df)
             - T*(p["cost_cian"]+p["transport_cian"]))
    nsr_h = ((T*l_au*rec_au_h*p["price_au"] + T*l_ag*rec_ag_h*p["price_ag"])*(1-df)
             - T*(p["cost_heap"]+p["transport_heap"]))

    def stats(arr):
        return {
            "mean": float(np.mean(arr)),
            "std":  float(np.std(arr)),
            "p5":   float(np.percentile(arr, 5)),
            "p10":  float(np.percentile(arr, 10)),
            "p50":  float(np.percentile(arr, 50)),
            "p90":  float(np.percentile(arr, 90)),
            "p95":  float(np.percentile(arr, 95)),
            "prob_positive": float(np.mean(arr > 0)),
        }

    correlations = {
        "Ley Au vs NSR Cian":  float(np.corrcoef(l_au, nsr_c)[0,1]),
        "Ley Ag vs NSR Cian":  float(np.corrcoef(l_ag, nsr_c)[0,1]),
        "Ley Au vs NSR Heap":  float(np.corrcoef(l_au, nsr_h)[0,1]),
        "Ley Ag vs NSR Heap":  float(np.corrcoef(l_ag, nsr_h)[0,1]),
        "Rec Au vs NSR Cian":  float(np.corrcoef(rec_au_c, nsr_c)[0,1]),
        "Rec Au vs NSR Heap":  float(np.corrcoef(rec_au_h, nsr_h)[0,1]),
    }

    return {
        "n": n, "l_au_sim": l_au, "l_ag_sim": l_ag,
        "nsr_cian": nsr_c, "nsr_heap": nsr_h,
        "stats_cian": stats(nsr_c), "stats_heap": stats(nsr_h), "stats_au": stats(l_au),
        "prob_above_co_cian": float(np.mean(l_au >= p["cutoff_cian"])),
        "prob_above_co_heap": float(np.mean(l_au >= p["cutoff_heap"])),
        "correlations": correlations,
    }


def calc_guard_balance(p, n_trucks):
    rng = np.random.default_rng(42)
    gau, gag = calc_corrected_grades(p["grade_au"], p["grade_ag"], p["fc_lab"])
    sig_ag   = p["grade_ag"] * 0.15
    df       = p["deduction"] / 100.0
    trucks   = []

    for k in range(n_trucks):
        t_k  = p["tonnage"] * rng.uniform(0.90, 1.10)
        l_au = max(0, rng.normal(gau, p["sigma_m"]))
        l_ag = max(0, rng.normal(gag, sig_ag))
        nsr_c, *_ = calc_nsr(t_k, l_au, l_ag,
                              p["rec_au_cian"]/100, p["rec_ag_cian"]/100,
                              p["price_au"], p["price_ag"], df,
                              p["cost_cian"], p["transport_cian"])
        nsr_h, *_ = calc_nsr(t_k, l_au, l_ag,
                              p["rec_au_heap"]/100, p["rec_ag_heap"]/100,
                              p["price_au"], p["price_ag"], df,
                              p["cost_heap"], p["transport_heap"])
        dest_opt, nsr_opt = calc_decision(nsr_c, nsr_h, l_au, p["cutoff_cian"], p["cutoff_heap"])

        if l_au >= p["cutoff_cian"]:
            dest_real = rng.choice(["CIANURACIÓN","HEAP LEACH","BOTADERO"], p=[0.55,0.35,0.10])
        elif l_au >= p["cutoff_heap"]:
            dest_real = rng.choice(["HEAP LEACH","BOTADERO"], p=[0.70,0.30])
        else:
            dest_real = "BOTADERO"

        if dest_real == "CIANURACIÓN":
            nsr_real = nsr_c; rec_au_r, rec_ag_r = p["rec_au_cian"]/100, p["rec_ag_cian"]/100
        elif dest_real == "HEAP LEACH":
            nsr_real = nsr_h; rec_au_r, rec_ag_r = p["rec_au_heap"]/100, p["rec_ag_heap"]/100
        else:
            nsr_real = 0.0; rec_au_r = rec_ag_r = 0.0

        trucks.append({
            "k": k+1, "tonnage": t_k, "grade_au": l_au, "grade_ag": l_ag,
            "nsr_opt": nsr_opt, "nsr_real": nsr_real,
            "dest_opt": dest_opt, "dest_real": dest_real,
            "mf_au": t_k * l_au * rec_au_r, "mf_ag": t_k * l_ag * rec_ag_r,
        })

    nsr_opt_t  = sum(t["nsr_opt"]  for t in trucks)
    nsr_real_t = sum(t["nsr_real"] for t in trucks)
    mf_au_t    = sum(t["mf_au"]    for t in trucks)
    mf_ag_t    = sum(t["mf_ag"]    for t in trucks)
    eff = (nsr_real_t / nsr_opt_t * 100) if nsr_opt_t > 0 else 0.0
    dc  = {"CIANURACIÓN": 0, "HEAP LEACH": 0, "BOTADERO": 0}
    for t in trucks:
        dc[t["dest_opt"]] += 1

    return {
        "trucks": trucks, "nsr_total_opt": nsr_opt_t, "nsr_total_real": nsr_real_t,
        "mf_au_total": mf_au_t, "mf_ag_total": mf_ag_t,
        "value_lost": nsr_opt_t - nsr_real_t, "efficiency": eff, "dest_counts": dc,
    }


# =============================================================================
# SECTION 2 — PLOT FUNCTIONS
# =============================================================================

def fig_nsr_histogram(mc):
    fig, axes = plt.subplots(1, 2, figsize=(13, 4.5))
    fig.suptitle("Distribución NSR Monte Carlo — Cianuración vs Heap Leach",
                 fontsize=12, fontweight="bold", color="white")
    for ax, key, label, color, skey in [
        (axes[0], "nsr_cian", "Cianuración", COLOR_CYAN, "stats_cian"),
        (axes[1], "nsr_heap", "Heap Leach",  COLOR_HEAP, "stats_heap"),
    ]:
        s = mc[skey]
        ax.hist(mc[key], bins=80, color=color, alpha=0.75, edgecolor="none", density=True)
        ax.axvline(s["mean"], color="white", lw=1.5, ls="--", label=f"Media: ${s['mean']:,.0f}")
        ax.axvline(s["p10"],  color=COLOR_YELLOW, lw=1.2, ls=":", label=f"P10: ${s['p10']:,.0f}")
        ax.axvline(s["p90"],  color=COLOR_GREEN,  lw=1.2, ls=":", label=f"P90: ${s['p90']:,.0f}")
        ax.axvline(0, color=COLOR_RED, lw=1.0, alpha=0.8, label="NSR=0")
        ax.set_xlabel("NSR ($)"); ax.set_ylabel("Densidad")
        ax.set_title(label, color=color); ax.legend(fontsize=8); ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return fig


def fig_cdf(mc):
    fig, axes = plt.subplots(1, 2, figsize=(13, 4.5))
    fig.suptitle("CDF del NSR", fontsize=12, fontweight="bold", color="white")
    for ax, key, label, color in [
        (axes[0], "nsr_cian", "Cianuración", COLOR_CYAN),
        (axes[1], "nsr_heap", "Heap Leach",  COLOR_HEAP),
    ]:
        data = np.sort(mc[key]); cp = np.arange(1, len(data)+1)/len(data)
        ax.plot(data, cp, color=color, lw=2)
        ax.axhline(0.10, color=COLOR_YELLOW, lw=1, ls="--", label="P10")
        ax.axhline(0.50, color="white",      lw=1, ls="--", label="P50")
        ax.axhline(0.90, color=COLOR_GREEN,  lw=1, ls="--", label="P90")
        ax.axvline(0, color=COLOR_RED, lw=1, alpha=0.7)
        ax.set_xlabel("NSR ($)"); ax.set_ylabel("Prob. acumulada")
        ax.set_title(label, color=color); ax.legend(fontsize=8); ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return fig


def fig_tornado(mc):
    corr  = mc["correlations"]; keys = list(corr.keys()); vals = [corr[k] for k in keys]
    order = np.argsort(np.abs(vals))[::-1]
    keys  = [keys[i] for i in order]; vals = [vals[i] for i in order]
    colors= [COLOR_GREEN if v >= 0 else COLOR_RED for v in vals]
    fig, ax = plt.subplots(figsize=(10, 4.5))
    bars = ax.barh(keys, vals, color=colors, alpha=0.85, edgecolor="none")
    ax.axvline(0, color="white", lw=1)
    for bar, val in zip(bars, vals):
        ax.text(val+(0.01 if val>=0 else -0.01), bar.get_y()+bar.get_height()/2,
                f"{val:.3f}", va="center", ha="left" if val>=0 else "right", fontsize=9, color="white")
    ax.set_xlabel("Correlación de Pearson"); ax.set_xlim(-1.1, 1.1)
    ax.set_title("Diagrama de Tornado — Sensibilidad", fontsize=12, fontweight="bold", color="white")
    ax.grid(True, axis="x", alpha=0.3)
    ax.legend(handles=[mpatches.Patch(color=COLOR_GREEN, label="Positiva"),
                       mpatches.Patch(color=COLOR_RED,   label="Negativa")], fontsize=9)
    fig.tight_layout()
    return fig


def fig_scatter(mc):
    fig, axes = plt.subplots(1, 2, figsize=(13, 4.5))
    fig.suptitle("Ley Au vs NSR por Ruta", fontsize=12, fontweight="bold", color="white")
    sl = slice(0, 2000)
    for ax, key, label, color in [
        (axes[0], "nsr_cian", "Cianuración", COLOR_CYAN),
        (axes[1], "nsr_heap", "Heap Leach",  COLOR_HEAP),
    ]:
        ax.scatter(mc["l_au_sim"][sl], mc[key][sl], alpha=0.25, s=4, color=color)
        ax.axhline(0, color=COLOR_RED, lw=1, ls="--")
        ax.set_xlabel("Ley Au (g/t)"); ax.set_ylabel("NSR ($)")
        ax.set_title(label, color=color); ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return fig


def fig_grade_dist(mc, p):
    gau = p["grade_au"] * p["fc_lab"]
    fig, axes = plt.subplots(1, 2, figsize=(13, 4.5))
    fig.suptitle("Distribución de Leyes Simuladas", fontsize=12, fontweight="bold", color="white")
    axes[0].hist(mc["l_au_sim"], bins=80, color=COLOR_CYAN, alpha=0.75, edgecolor="none", density=True)
    axes[0].axvline(p["cutoff_cian"], color=COLOR_CYAN, lw=1.5, ls="--", label=f"CO_cian={p['cutoff_cian']:.2f}")
    axes[0].axvline(p["cutoff_heap"], color=COLOR_HEAP, lw=1.5, ls=":",  label=f"CO_heap={p['cutoff_heap']:.2f}")
    axes[0].axvline(gau, color="white", lw=1.5, ls="-.", label=f"L_Au corr={gau:.3f}")
    axes[0].set_xlabel("Ley Au (g/t)"); axes[0].set_ylabel("Densidad")
    axes[0].set_title("Ley Au simulada", color=COLOR_CYAN); axes[0].legend(fontsize=8); axes[0].grid(True, alpha=0.3)
    axes[1].hist(mc["l_ag_sim"], bins=80, color=COLOR_HEAP, alpha=0.75, edgecolor="none", density=True)
    axes[1].set_xlabel("Ley Ag (g/t)"); axes[1].set_ylabel("Densidad")
    axes[1].set_title("Ley Ag simulada", color=COLOR_HEAP); axes[1].grid(True, alpha=0.3)
    fig.tight_layout()
    return fig


def fig_nsr_comparison(det):
    fig, ax = plt.subplots(figsize=(8, 4.5))
    routes = ["Cianuración", "Heap Leach", "Botadero"]
    values = [det["nsr_cian"], det["nsr_heap"], 0.0]
    colors = [COLOR_CYAN if det["nsr_cian"]>0 else COLOR_RED,
              COLOR_HEAP if det["nsr_heap"]>0 else COLOR_RED, COLOR_DUMP]
    bars = ax.bar(routes, values, color=colors, alpha=0.85, edgecolor="none", width=0.5)
    ax.axhline(0, color="white", lw=1, ls="--")
    for bar, val in zip(bars, values):
        ax.text(bar.get_x()+bar.get_width()/2, val+(10 if val>=0 else -10),
                f"${val:,.0f}", ha="center", va="bottom" if val>=0 else "top",
                fontsize=11, fontweight="bold", color="white")
    ax.set_ylabel("NSR Determinístico ($)")
    ax.set_title("Comparación NSR por Ruta", fontsize=12, fontweight="bold", color="white")
    ax.grid(True, axis="y", alpha=0.3); fig.tight_layout()
    return fig


def fig_guard_balance(balance):
    trucks = balance["trucks"]
    k_vals  = [t["k"] for t in trucks]
    nsr_opt = [t["nsr_opt"] for t in trucks]; nsr_real = [t["nsr_real"] for t in trucks]
    fig = plt.figure(figsize=(14, 8))
    gs  = GridSpec(2, 2, figure=fig, hspace=0.4, wspace=0.35)
    fig.suptitle("Balance de Guardia Completa", fontsize=13, fontweight="bold", color="white")

    ax1 = fig.add_subplot(gs[0, :]); ax1.set_facecolor("#1E2130")
    ax1.plot(k_vals, nsr_opt,  color=COLOR_GREEN,  lw=2, label="NSR Óptimo", marker="o", ms=3)
    ax1.plot(k_vals, nsr_real, color=COLOR_YELLOW, lw=2, label="NSR Real",   marker="s", ms=3)
    ax1.fill_between(k_vals, nsr_opt, nsr_real,
                     where=[o>r for o,r in zip(nsr_opt,nsr_real)],
                     alpha=0.2, color=COLOR_RED, label="Valor no capturado")
    ax1.axhline(0, color="white", lw=0.8, ls="--")
    ax1.set_xlabel("N° Volquete"); ax1.set_ylabel("NSR ($)")
    ax1.set_title("NSR Óptimo vs Real"); ax1.legend(fontsize=9); ax1.grid(True, alpha=0.3)

    ax2 = fig.add_subplot(gs[1, 0]); ax2.set_facecolor("#1E2130")
    dc   = balance["dest_counts"]
    bars = ax2.bar(dc.keys(), dc.values(), color=[COLOR_CYAN,COLOR_HEAP,COLOR_DUMP], alpha=0.85, edgecolor="none")
    for bar, v in zip(bars, dc.values()):
        ax2.text(bar.get_x()+bar.get_width()/2, v+0.3, str(v), ha="center", va="bottom", color="white", fontsize=10)
    ax2.set_ylabel("N° Volquetes"); ax2.set_title("Destino Óptimo"); ax2.grid(True, axis="y", alpha=0.3)

    ax3 = fig.add_subplot(gs[1, 1]); ax3.set_facecolor("#1E2130")
    vals3 = [balance["nsr_total_opt"], balance["nsr_total_real"], balance["value_lost"]]
    bars3 = ax3.bar(["NSR Óptimo","NSR Real","Valor perdido"], vals3,
                    color=[COLOR_GREEN,COLOR_YELLOW,COLOR_RED], alpha=0.85, edgecolor="none")
    for bar, v in zip(bars3, vals3):
        ax3.text(bar.get_x()+bar.get_width()/2, v+max(vals3)*0.01,
                 f"${v:,.0f}", ha="center", va="bottom", color="white", fontsize=9)
    ax3.set_ylabel("NSR Total ($)"); ax3.set_title(f"Eficiencia: {balance['efficiency']:.1f}%")
    ax3.grid(True, axis="y", alpha=0.3)
    return fig


def fig_polar(mc):
    corr   = mc["correlations"]; labels = list(corr.keys()); vals = [abs(corr[k]) for k in labels]
    n      = len(labels); angles = np.linspace(0, 2*np.pi, n, endpoint=False).tolist()
    vals_p = vals + vals[:1]; ang_p = angles + angles[:1]
    fig, ax = plt.subplots(figsize=(7, 7), subplot_kw={"polar": True})
    ax.set_facecolor("#1E2130"); fig.patch.set_facecolor("#0E1117")
    ax.plot(ang_p, vals_p, color=COLOR_ACCENT, lw=2)
    ax.fill(ang_p, vals_p, color=COLOR_ACCENT, alpha=0.3)
    ax.set_xticks(angles); ax.set_xticklabels(labels, fontsize=8, color="white")
    ax.set_ylim(0, 1.05)
    ax.set_title("Diagrama Polar de Sensibilidad", fontsize=11, fontweight="bold", color="white", pad=20)
    ax.grid(color="#3A3F58", alpha=0.5)
    return fig


def fig_boxplot(mc):
    fig, ax = plt.subplots(figsize=(8, 4.5))
    bp = ax.boxplot([mc["nsr_cian"], mc["nsr_heap"]], labels=["Cianuración","Heap Leach"],
                    patch_artist=True,
                    boxprops=dict(facecolor="#1E2130", color="white"),
                    medianprops=dict(color=COLOR_YELLOW, lw=2),
                    whiskerprops=dict(color="white"), capprops=dict(color="white"),
                    flierprops=dict(marker="o", color=COLOR_ACCENT, alpha=0.3, ms=2))
    bp["boxes"][0].set_facecolor(COLOR_CYAN+"44"); bp["boxes"][1].set_facecolor(COLOR_HEAP+"44")
    ax.axhline(0, color=COLOR_RED, lw=1.2, ls="--", label="NSR=0")
    ax.set_ylabel("NSR ($)"); ax.legend(fontsize=9)
    ax.set_title("Boxplot Comparativo NSR", fontsize=12, fontweight="bold", color="white")
    ax.grid(True, axis="y", alpha=0.3); fig.tight_layout()
    return fig


# =============================================================================
# SECTION 3 — EXCEL EXPORT
# =============================================================================

def _hfill(color_hex):
    return PatternFill(fill_type="solid", fgColor=color_hex)


def export_excel_bytes(p, det, mc, balance):
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    def header_row(ws, row, texts, col_start=1):
        for col, h in enumerate(texts, col_start):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = _hfill(EXCEL_HEADER_FILL)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

    def title_cell(ws, text, merge="A1:E1"):
        ws.merge_cells(merge)
        c = ws[merge.split(":")[0]]; c.value = text
        c.fill = _hfill(EXCEL_HEADER_FILL)
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.alignment = Alignment(horizontal="center")

    # Sheet 1
    ws1 = wb.create_sheet("1_Datos_Entrada")
    title_cell(ws1, "MINEBALANCE — DATOS DE ENTRADA", "A1:E1")
    header_row(ws1, 2, ["Parámetro","Valor","Unidad","Rango típico","CoV MC (%)"])
    rows_data = [
        ("Labor",            p["labor_id"],      "texto", "Nv-380, Nv-420...", "N/A"),
        ("Tonelaje T",       p["tonnage"],        "t",     "5–30 t",            "N/A"),
        ("Ley Au",           p["grade_au"],       "g/t",   "0.5–12.0",          "N/A"),
        ("Ley Ag",           p["grade_ag"],       "g/t",   "5–200",             "N/A (15%)"),
        ("Error σ_m",        p["sigma_m"],        "g/t",   "10–50% L̂_Au",      "N/A"),
        ("FC_lab",           p["fc_lab"],         "adim.", "0.70–1.20",         "N/A"),
        ("Precio Au",        p["price_au"],       "$/g",   "33–62",             "N/A"),
        ("Precio Ag",        p["price_ag"],       "$/g",   "0.30–0.65",         "N/A"),
        ("Deducción",        p["deduction"],      "%",     "3–10",              "N/A"),
        ("Rec Au Cian",      p["rec_au_cian"],    "%",     "88–96",             "3%"),
        ("Rec Ag Cian",      p["rec_ag_cian"],    "%",     "70–90",             "4%"),
        ("Costo Cian",       p["cost_cian"],      "$/t",   "120–250",           "N/A"),
        ("Transp. Cian",     p["transport_cian"], "$/t",   "5–30",              "N/A"),
        ("Cut-off Cian",     p["cutoff_cian"],    "g/t",   "1.0–4.0",           "N/A"),
        ("Rec Au Heap",      p["rec_au_heap"],    "%",     "45–75",             "5%"),
        ("Rec Ag Heap",      p["rec_ag_heap"],    "%",     "20–55",             "6%"),
        ("Costo Heap",       p["cost_heap"],      "$/t",   "35–100",            "N/A"),
        ("Transp. Heap",     p["transport_heap"], "$/t",   "3–15",              "N/A"),
        ("Cut-off Heap",     p["cutoff_heap"],    "g/t",   "0.4–2.0",           "N/A"),
        ("Simulaciones",     p["n_simulations"],  "iter.", "5,000–50,000",      "N/A"),
    ]
    for ri, (param, val, unit, typical, cov) in enumerate(rows_data, 3):
        ws1.cell(row=ri,column=1,value=param); ws1.cell(row=ri,column=2,value=val)
        ws1.cell(row=ri,column=3,value=unit);  ws1.cell(row=ri,column=4,value=typical)
        ws1.cell(row=ri,column=5,value=cov)
    for w,col in zip([28,14,10,24,12],range(1,6)):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # Sheet 2
    ws2 = wb.create_sheet("2_Resultados_NSR")
    title_cell(ws2, "RESULTADOS NSR — DETERMINÍSTICO Y MONTE CARLO", "A1:D1")
    header_row(ws2, 2, ["Indicador","Cianuración","Heap Leach","Unidad"])
    gau = p["grade_au"] * p["fc_lab"]
    sc = mc["stats_cian"]; sh = mc["stats_heap"]
    dest, _ = calc_decision(det["nsr_cian"], det["nsr_heap"], gau, p["cutoff_cian"], p["cutoff_heap"])
    res_rows = [
        ("Ley Au corregida", f"{gau:.4f}", f"{gau:.4f}", "g/t"),
        ("NSR Determinístico", f"${det['nsr_cian']:,.2f}", f"${det['nsr_heap']:,.2f}", "$"),
        ("E[NSR] Monte Carlo", f"${sc['mean']:,.1f}", f"${sh['mean']:,.1f}", "$"),
        ("P10",  f"${sc['p10']:,.0f}", f"${sh['p10']:,.0f}", "$"),
        ("P50",  f"${sc['p50']:,.0f}", f"${sh['p50']:,.0f}", "$"),
        ("P90",  f"${sc['p90']:,.0f}", f"${sh['p90']:,.0f}", "$"),
        ("P(NSR>0)", f"{sc['prob_positive']*100:.1f}%", f"{sh['prob_positive']*100:.1f}%", "%"),
        ("DECISIÓN", dest, dest, "—"),
    ]
    for ri,(ind,vc,vh,unit) in enumerate(res_rows, 3):
        ws2.cell(row=ri,column=1,value=ind).font = Font(bold=True)
        ws2.cell(row=ri,column=2,value=vc); ws2.cell(row=ri,column=3,value=vh)
        ws2.cell(row=ri,column=4,value=unit)
        if ind == "DECISIÓN":
            fc = EXCEL_GREEN_FILL if dest=="CIANURACIÓN" else EXCEL_YELLOW_FILL if dest=="HEAP LEACH" else EXCEL_RED_FILL
            for col in [2,3]:
                ws2.cell(row=ri,column=col).fill = _hfill(fc)
                ws2.cell(row=ri,column=col).font = Font(bold=True)
    for w,col in zip([28,20,20,12],range(1,5)):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # Sheet 3 — MC Sample
    ws3 = wb.create_sheet("3_Muestra_MC")
    title_cell(ws3, "MUESTRA MONTE CARLO — 500 FILAS", "A1:E1")
    header_row(ws3, 2, ["Iteración","Ley Au (g/t)","Ley Ag (g/t)","NSR Cian ($)","NSR Heap ($)"])
    for i in range(min(500, mc["n"])):
        ws3.cell(row=i+3,column=1,value=i+1)
        ws3.cell(row=i+3,column=2,value=round(float(mc["l_au_sim"][i]),4))
        ws3.cell(row=i+3,column=3,value=round(float(mc["l_ag_sim"][i]),2))
        ws3.cell(row=i+3,column=4,value=round(float(mc["nsr_cian"][i]),2))
        ws3.cell(row=i+3,column=5,value=round(float(mc["nsr_heap"][i]),2))
    for w,col in zip([12,16,16,18,18],range(1,6)):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # Sheet 4 — Guard balance
    ws4 = wb.create_sheet("4_Balance_Guardia")
    title_cell(ws4, f"BALANCE DE GUARDIA — {len(balance['trucks'])} VOLQUETES", "A1:G1")
    header_row(ws4, 2, ["Vol.","Tonelaje","Ley Au","NSR Óptimo","NSR Real","Dest. Óptimo","Dest. Real"])
    for ri,t in enumerate(balance["trucks"],3):
        ws4.cell(row=ri,column=1,value=t["k"]); ws4.cell(row=ri,column=2,value=round(t["tonnage"],2))
        ws4.cell(row=ri,column=3,value=round(t["grade_au"],4)); ws4.cell(row=ri,column=4,value=round(t["nsr_opt"],2))
        ws4.cell(row=ri,column=5,value=round(t["nsr_real"],2)); ws4.cell(row=ri,column=6,value=t["dest_opt"])
        ws4.cell(row=ri,column=7,value=t["dest_real"])
    for w,col in zip([10,14,14,18,18,18,18],range(1,8)):
        ws4.column_dimensions[get_column_letter(col)].width = w

    # Sheet 5 — Correlations
    ws5 = wb.create_sheet("5_Correlaciones")
    title_cell(ws5, "CORRELACIONES DE PEARSON", "A1:C1")
    header_row(ws5, 2, ["Variable","Correlación","Interpretación"])
    for ri,(var,val) in enumerate(mc["correlations"].items(),3):
        ws5.cell(row=ri,column=1,value=var); ws5.cell(row=ri,column=2,value=round(val,4))
        ws5.cell(row=ri,column=3,value=(
            "Fuerte positiva" if val>=0.7 else "Fuerte negativa" if val<=-0.7
            else "Moderada" if abs(val)>=0.4 else "Débil"))
    for w,col in zip([38,18,26],range(1,4)):
        ws5.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# =============================================================================
# SECTION 4 — STREAMLIT UI
# =============================================================================

def main():
    st.markdown("""
    <div style='text-align:center; padding:1rem 0 0.5rem 0;'>
        <h1 style='color:#00C6FF; font-size:2.4rem; margin-bottom:0;'>⛏️ MineBalance</h1>
        <p style='color:#C9D1E0; font-size:1.05rem; margin-top:0.3rem;'>
            Sistema de Optimización de Destino de Mineral · NSR · Monte Carlo · FC Adaptativo
        </p>
        <p style='color:#7A8099; font-size:0.85rem;'>
            Universidad Nacional del Altiplano &nbsp;|&nbsp; Ingeniería de Minas &nbsp;|&nbsp; 9no Semestre 2026
            &nbsp;|&nbsp; <b>Machaca Espinoza, Jkacson Ruso</b>
        </p>
    </div>
    <hr style='border-color:#2E3450;'>
    """, unsafe_allow_html=True)

    # ── Sidebar ─────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## ⚙️ Parámetros de Entrada")

        st.markdown("### 📍 Sección A — Volquete y Muestreo")
        labor_id   = st.text_input("Identificador de labor", value="Nv-420")
        tonnage    = st.number_input("Tonelaje por viaje T (t)",       1.0, 100.0, 20.0, 0.5)
        grade_au   = st.number_input("Ley Au estimada L̂_Au (g/t)",    0.1,  30.0,  3.5, 0.1)
        grade_ag   = st.number_input("Ley Ag estimada L̂_Ag (g/t)",    1.0, 500.0, 45.0, 1.0)
        sigma_m    = st.number_input("Error muestreo σ_m (g/t)",      0.01,   5.0,  0.70, 0.05)
        fc_lab     = st.number_input("Factor corrección FC_lab",       0.50,  1.50,  0.91, 0.01)

        st.markdown("### 💰 Sección B — Precios")
        price_au   = st.number_input("Precio Au P_Au ($/g)",          10.0, 100.0, 39.0, 0.5)
        price_ag   = st.number_input("Precio Ag P_Ag ($/g)",          0.10,   2.0,  0.48, 0.01)
        deduction  = st.number_input("Deducción fundición D_f (%)",    1.0,  15.0,  5.0, 0.5)

        st.markdown("### 🏭 Sección C — Cianuración")
        rec_au_cian    = st.number_input("Rec. Au Cian (%)",      50.0, 99.0, 92.0, 0.5)
        rec_ag_cian    = st.number_input("Rec. Ag Cian (%)",      30.0, 99.0, 78.0, 0.5)
        cost_cian      = st.number_input("Costo proceso Cian ($/t)", 50.0, 500.0, 185.0, 5.0)
        transport_cian = st.number_input("Transporte Cian ($/t)",   2.0,  50.0,  12.0, 1.0)
        cutoff_cian    = st.number_input("Cut-off Cian (g/t Au)",   0.5,   8.0,   2.50, 0.1)

        st.markdown("### ⛰️ Sección D — Heap Leach")
        rec_au_heap    = st.number_input("Rec. Au Heap (%)",      20.0, 90.0, 65.0, 0.5)
        rec_ag_heap    = st.number_input("Rec. Ag Heap (%)",      10.0, 80.0, 38.0, 0.5)
        cost_heap      = st.number_input("Costo proceso Heap ($/t)", 10.0, 200.0, 62.0, 5.0)
        transport_heap = st.number_input("Transporte Heap ($/t)",   1.0,  30.0,   7.0, 1.0)
        cutoff_heap    = st.number_input("Cut-off Heap (g/t Au)",   0.1,   5.0,   1.20, 0.1)

        st.markdown("### 🎲 Sección E — Simulación")
        n_simulations = st.select_slider("Simulaciones Monte Carlo",
                                         options=[1000,5000,10000,20000,50000], value=10000)
        n_trucks      = st.slider("Volquetes en guardia", 5, 120, 40)

        run_btn = st.button("▶  EJECUTAR ANÁLISIS", type="primary", use_container_width=True)

    p = {
        "labor_id": labor_id, "tonnage": tonnage,
        "grade_au": grade_au, "grade_ag": grade_ag,
        "sigma_m": sigma_m, "fc_lab": fc_lab,
        "price_au": price_au, "price_ag": price_ag, "deduction": deduction,
        "rec_au_cian": rec_au_cian, "rec_ag_cian": rec_ag_cian,
        "cost_cian": cost_cian, "transport_cian": transport_cian, "cutoff_cian": cutoff_cian,
        "rec_au_heap": rec_au_heap, "rec_ag_heap": rec_ag_heap,
        "cost_heap": cost_heap, "transport_heap": transport_heap, "cutoff_heap": cutoff_heap,
        "n_simulations": n_simulations,
    }

    if not run_btn:
        st.info("👈  Configura los parámetros en el panel izquierdo y presiona **EJECUTAR ANÁLISIS**.")
        st.markdown("""
        <div style='display:flex; gap:1rem; flex-wrap:wrap; margin-top:1.5rem;'>
        <div style='flex:1;min-width:200px;background:#1E2130;border-radius:10px;padding:1rem;border-left:4px solid #00C6FF;'>
            <b style='color:#00C6FF;'>NSR Determinístico</b><br>
            <span style='color:#C9D1E0;font-size:.9rem;'>Valor económico neto por volquete según ruta de procesamiento con parámetros fijos.</span>
        </div>
        <div style='flex:1;min-width:200px;background:#1E2130;border-radius:10px;padding:1rem;border-left:4px solid #7C4DFF;'>
            <b style='color:#7C4DFF;'>Monte Carlo</b><br>
            <span style='color:#C9D1E0;font-size:.9rem;'>Simulación probabilística de la incertidumbre en la ley del muestreo de frente.</span>
        </div>
        <div style='flex:1;min-width:200px;background:#1E2130;border-radius:10px;padding:1rem;border-left:4px solid #FFB347;'>
            <b style='color:#FFB347;'>Balance de Guardia</b><br>
            <span style='color:#C9D1E0;font-size:.9rem;'>Eficiencia de las decisiones tomadas durante el turno completo.</span>
        </div>
        </div>
        """, unsafe_allow_html=True)
        return

    # ── Run calculations ─────────────────────────────────────────────────────
    with st.spinner("⚙️ Calculando..."):
        gau, gag = calc_corrected_grades(p["grade_au"], p["grade_ag"], p["fc_lab"])
        df = p["deduction"] / 100.0
        nsr_c, gross_au_c, gross_ag_c, income_c, cost_c = calc_nsr(
            p["tonnage"], gau, gag,
            p["rec_au_cian"]/100, p["rec_ag_cian"]/100,
            p["price_au"], p["price_ag"], df,
            p["cost_cian"], p["transport_cian"])
        nsr_h, gross_au_h, gross_ag_h, income_h, cost_h = calc_nsr(
            p["tonnage"], gau, gag,
            p["rec_au_heap"]/100, p["rec_ag_heap"]/100,
            p["price_au"], p["price_ag"], df,
            p["cost_heap"], p["transport_heap"])
        det = {"nsr_cian": nsr_c, "nsr_heap": nsr_h}
        mc  = monte_carlo(p)
        balance = calc_guard_balance(p, n_trucks)

    sc = mc["stats_cian"]; sh = mc["stats_heap"]
    dest, _ = calc_decision(nsr_c, nsr_h, gau, p["cutoff_cian"], p["cutoff_heap"])

    # ── Decision banner ───────────────────────────────────────────────────────
    color_map = {"CIANURACIÓN": "#00C6FF", "HEAP LEACH": "#FFB347", "BOTADERO": "#FF6B6B"}
    icon_map  = {"CIANURACIÓN": "🔵", "HEAP LEACH": "🟠", "BOTADERO": "🔴"}
    dc = color_map[dest]
    st.markdown(f"""
    <div style='background:#1E2130;border-radius:12px;padding:1.2rem 2rem;
                border-left:6px solid {dc};margin-bottom:1rem;'>
        <h2 style='color:{dc};margin:0;'>{icon_map[dest]} DECISIÓN: {dest}</h2>
        <p style='color:#C9D1E0;margin:.3rem 0 0 0;font-size:.95rem;'>
            Labor: <b>{p['labor_id']}</b> &nbsp;|&nbsp;
            L_Au = <b>{p['grade_au']:.3f} g/t</b> &nbsp;|&nbsp;
            FC_lab = <b>{p['fc_lab']:.4f}</b> &nbsp;→&nbsp;
            L_Au corregida = <b>{gau:.4f} g/t</b>
        </p>
    </div>
    """, unsafe_allow_html=True)

    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("NSR Det. Cian ($)",  f"${nsr_c:,.0f}")
    c2.metric("NSR Det. Heap ($)",  f"${nsr_h:,.0f}")
    c3.metric("E[NSR] Heap MC ($)", f"${sh['mean']:,.0f}")
    c4.metric("P(NSR Heap > 0)",    f"{sh['prob_positive']*100:.1f}%")
    c5.metric("Eficiencia Guardia", f"{balance['efficiency']:.1f}%")
    st.markdown("---")

    # ── Tabs ──────────────────────────────────────────────────────────────────
    tab1,tab2,tab3,tab4,tab5,tab6 = st.tabs([
        "📊 NSR Determinístico","🎲 Monte Carlo",
        "🔍 Sensibilidad","🚛 Balance Guardia",
        "📋 Resumen","📥 Descargar Excel",
    ])

    with tab1:
        st.subheader("Cálculo Paso a Paso — NSR Determinístico")
        ca, cb = st.columns(2)
        with ca:
            st.markdown(f"""**Cianuración (CIL/CIP)**
- Ingreso bruto Au: **${gross_au_c:,.2f}**
- Ingreso bruto Ag: **${gross_ag_c:,.2f}**
- Ingreso neto ({p['deduction']:.1f}% deducción): **${income_c:,.2f}**
- Costo total: **${cost_c:,.2f}**
- **NSR = ${nsr_c:,.2f}**""")
        with cb:
            st.markdown(f"""**Heap Leach**
- Ingreso bruto Au: **${gross_au_h:,.2f}**
- Ingreso bruto Ag: **${gross_ag_h:,.2f}**
- Ingreso neto ({p['deduction']:.1f}% deducción): **${income_h:,.2f}**
- Costo total: **${cost_h:,.2f}**
- **NSR = ${nsr_h:,.2f}**""")
        st.pyplot(fig_nsr_comparison(det))

    with tab2:
        st.subheader(f"Simulación Monte Carlo — N = {mc['n']:,} iteraciones")
        ca, cb = st.columns(2)
        with ca:
            st.markdown("**Cianuración**")
            st.dataframe({"Estadístico":["Media","Desv. Std","P10","P50","P90","P(NSR>0)"],
                          "Valor":[f"${sc['mean']:,.1f}",f"${sc['std']:,.1f}",
                                   f"${sc['p10']:,.0f}",f"${sc['p50']:,.0f}",f"${sc['p90']:,.0f}",
                                   f"{sc['prob_positive']*100:.1f}%"]},
                         use_container_width=True, hide_index=True)
        with cb:
            st.markdown("**Heap Leach**")
            st.dataframe({"Estadístico":["Media","Desv. Std","P10","P50","P90","P(NSR>0)"],
                          "Valor":[f"${sh['mean']:,.1f}",f"${sh['std']:,.1f}",
                                   f"${sh['p10']:,.0f}",f"${sh['p50']:,.0f}",f"${sh['p90']:,.0f}",
                                   f"{sh['prob_positive']*100:.1f}%"]},
                         use_container_width=True, hide_index=True)
        st.pyplot(fig_nsr_histogram(mc))
        st.pyplot(fig_cdf(mc))
        st.pyplot(fig_grade_dist(mc, p))
        st.pyplot(fig_boxplot(mc))
        st.pyplot(fig_scatter(mc))

    with tab3:
        st.subheader("Análisis de Sensibilidad")
        st.pyplot(fig_tornado(mc))
        st.pyplot(fig_polar(mc))
        st.markdown("**Correlaciones de Pearson**")
        st.dataframe({
            "Variable": list(mc["correlations"].keys()),
            "Correlación": [f"{v:.4f}" for v in mc["correlations"].values()],
            "Interpretación": [
                "Fuerte positiva" if v>=0.7 else "Fuerte negativa" if v<=-0.7
                else "Moderada" if abs(v)>=0.4 else "Débil"
                for v in mc["correlations"].values()
            ],
        }, use_container_width=True, hide_index=True)

    with tab4:
        st.subheader(f"Balance de Guardia — {n_trucks} Volquetes")
        ca,cb,cc,cd = st.columns(4)
        ca.metric("NSR Óptimo Total",   f"${balance['nsr_total_opt']:,.0f}")
        cb.metric("NSR Real Total",     f"${balance['nsr_total_real']:,.0f}")
        cc.metric("Valor no capturado", f"${balance['value_lost']:,.0f}")
        cd.metric("Metal Au (guardia)", f"{balance['mf_au_total']:.1f} g = {balance['mf_au_total']/31.1035:.2f} oz")
        st.pyplot(fig_guard_balance(balance))
        with st.expander("📋 Ver tabla de volquetes"):
            import pandas as pd
            st.dataframe(pd.DataFrame([{
                "Vol.": t["k"], "Tonelaje": f"{t['tonnage']:.1f} t",
                "Ley Au": f"{t['grade_au']:.4f} g/t",
                "NSR Óptimo": f"${t['nsr_opt']:,.0f}",
                "NSR Real": f"${t['nsr_real']:,.0f}",
                "Dest. Óptimo": t["dest_opt"], "Dest. Real": t["dest_real"],
            } for t in balance["trucks"]]), use_container_width=True, hide_index=True)

    with tab5:
        st.subheader("📋 Resumen Ejecutivo")
        st.markdown(f"""
| Indicador | Valor |
|---|---|
| **Labor analizada** | {p['labor_id']} |
| **Ley Au estimada** | {p['grade_au']:.4f} g/t |
| **Ley Au corregida** (FC={p['fc_lab']:.4f}) | {gau:.4f} g/t |
| **NSR Cianuración** | ${nsr_c:,.2f} |
| **NSR Heap Leach** | ${nsr_h:,.2f} |
| **E[NSR] Heap MC** | ${sh['mean']:,.1f} |
| **P10 / P90 Heap MC** | ${sh['p10']:,.0f} / ${sh['p90']:,.0f} |
| **P(NSR Heap > 0)** | {sh['prob_positive']*100:.1f}% |
| **P(Ley ≥ CO_cian)** | {mc['prob_above_co_cian']*100:.1f}% |
| **P(Ley ≥ CO_heap)** | {mc['prob_above_co_heap']*100:.1f}% |
| **DECISIÓN** | **{dest}** |
| **NSR Óptimo guardia** | ${balance['nsr_total_opt']:,.0f} |
| **Eficiencia decisiones** | {balance['efficiency']:.1f}% |
| **Metal Au guardia** | {balance['mf_au_total']:.1f} g = {balance['mf_au_total']/31.1035:.2f} oz |
| **Metal Ag guardia** | {balance['mf_ag_total']:.1f} g = {balance['mf_ag_total']/31.1035:.2f} oz |
        """)

    with tab6:
        st.subheader("📥 Descargar Resultados en Excel")
        st.info("El archivo Excel contiene 5 hojas: Datos de entrada, Resultados NSR, Muestra Monte Carlo, Balance de Guardia y Correlaciones.")
        with st.spinner("Generando Excel..."):
            excel_bytes = export_excel_bytes(p, det, mc, balance)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="⬇️  Descargar MineBalance_Resultados.xlsx",
            data=excel_bytes,
            file_name=f"MineBalance_{p['labor_id']}_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.markdown("---")
        st.markdown("**Universidad Nacional del Altiplano | Ingeniería de Minas | 9no Semestre 2026**")
        st.markdown("*Autor: Machaca Espinoza, Jkacson Ruso*")


if __name__ == "__main__":
    main()
